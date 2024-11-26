import time
import pywhatkit as kit
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

# File Paths
excel_file = "contacts.xlsx"  # Your Excel file containing contact numbers
brochure_path = "path/to/brochure.pdf"  # Path to your PDF brochure
image1_path = "path/to/image1.jpg"  # Path to the first image
image2_path = "path/to/image2.jpg"  # Path to the second image

# Read contacts from Excel
def get_contacts(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    contacts = [row[0].value for row in sheet.iter_rows(min_row=2)]  # Assuming numbers start from row 2
    return contacts

# Send Greeting, PDF, and Images
def send_whatsapp_messages(contacts):
    # Load WhatsApp Web
    driver = webdriver.Chrome()  # Ensure you have the ChromeDriver set up
    driver.get("https://web.whatsapp.com")
    print("Scan the QR code to log in to WhatsApp Web...")
    time.sleep(20)  # Time to scan the QR code and login
    
    for contact in contacts:
        print(f"Processing contact: {contact}")
        
        # Step 1: Send Greeting
        greeting_message = "Hello! Check out our new product."
        kit.sendwhatmsg_instantly(contact, greeting_message, wait_time=10, tab_close=True)
        time.sleep(5)  # Wait for the message to be sent

        # Step 2: Send PDF
        driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(brochure_path)
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[contains(text(), 'Send')]").click()
        time.sleep(5)  # Wait for the file to be sent

        # Step 3: Send First Image
        driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(image1_path)
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[contains(text(), 'Send')]").click()
        time.sleep(5)

        # Step 4: Send Second Image
        driver.find_element(By.CSS_SELECTOR, "span[data-icon='clip']").click()
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(image2_path)
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[contains(text(), 'Send')]").click()
        time.sleep(5)

    print("Messages sent to all contacts!")
    driver.quit()

# Main Program
if __name__ == "__main__":
    contacts_list = get_contacts(excel_file)
    send_whatsapp_messages(contacts_list)
